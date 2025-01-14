import openpyxl
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
import json
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import openai

openai.api_key = "change me"

def collect_social_media_accounts(websites):
    """
    Processes a list of websites, fetches social media accounts for each, and
    returns aggregated lists for each platform.

    Args:
        websites (str): A string containing websites separated by newlines.

    Returns:
        dict: A dictionary containing lists of social media accounts grouped by platform.
    """
    if websites == "غير موجود":
        return {
            "facebook": [],
            "instagram": [],
            "twitter": [],
            "linkedin": [],
            "snapchat": [],
            "other": []
        }

    # Split websites, remove duplicates, and clean up entries
    unique_websites = list(set(websites.split('\n')))
    
    # Initialize aggregated results
    aggregated_accounts = {
        "facebook": [],
        "instagram": [],
        "twitter": [],
        "linkedin": [],
        "snapchat": [],
        "other": []
    }

    for website in unique_websites:
        if website is None:
            continue
        try:
            print(f"Fetching social media for: {website}")
            response = get_social_media(website)
            try:
                response = json.loads(response)
            except:
                response = {}
            
            if "social_media" in response:
                social_media = response["social_media"]
                for platform, account in social_media.items():
                    if account and platform in aggregated_accounts:
                        aggregated_accounts[platform].append(account)
            else:
                pass
        
        except Exception as e:
            pass

    # Remove duplicates from each list
    for platform in aggregated_accounts:
        aggregated_accounts[platform] = list(set(aggregated_accounts[platform]))

    return aggregated_accounts


# Example Usage


def read_company_names(file_path):
    """Reads company names from the Excel file."""
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    company_names = []
    for row in sheet.iter_rows(values_only=True):
        if row[0] == "Company Name":
            continue
        company_names.append(row[0])
    return company_names

def setup_driver():
    """Sets up and returns the WebDriver."""
    chrome_options = Options()
    #chrome_options.add_argument("--headless")  # Uncomment for headless mode
    chrome_options.add_argument("--disable-gpu")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    driver = webdriver.Chrome(options=chrome_options)
    return driver

def search_google_maps(driver, query):
    """Searches for the query on Google Maps."""
    driver.get("https://www.google.com/maps")
    search_box = driver.find_element(By.NAME, "q")
    search_box.clear()
    search_box.send_keys(f'"{query}"')
    search_box.send_keys(Keys.RETURN)
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.CSS_SELECTOR, "div[role='feed']")))

def extract_related_business(driver):
    """Extracts related business data from the current page."""
    related_business = {}
    items = driver.find_elements(By.CSS_SELECTOR, "[data-item-id]")
    for item in items:
        item_id = item.get_attribute('data-item-id')
        text = item.text.strip()
        if "address" in item_id:
            related_business["address"] = text.split('\n')[1]
        elif "phone:tel" in item_id:
            related_business["phone"] = text.split('\n')[1]
        elif "authority" in item_id:
            if "instagram" in text:
                related_business['url'] = None
            else:
                related_business["url"] = text.split('\n')[1]
    return related_business

def save_to_json(data, file_name):
    """Saves the scraped data to a JSON file."""
    with open(file_name, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=4)

def save_to_xls(data, file_name):
    """Saves the scraped data to a Excel file."""
    wb = openpyxl.load_workbook(file_name)
    ws = wb.active
    pass

def get_social_media(url):
    """
    Queries OpenAI to get social media accounts for the given URL.
    """
    prompt = f"""
    Please provide official social media accounts (Twitter, Facebook, LinkedIn, Instagram, Snapchat, Other) for the company with the URL "{url}".
    If no accounts are found, respond in the following JSON format:
    {{
        "company_url": "{url}",
        "social_media": {{
            "twitter": null,
            "facebook": null,
            "linkedin": null,
            "instagram": null,
            "snapchat": null,
            "other": null
        }}
    }}
    """
    try:
        if url is not None:
            response = openai.ChatCompletion.create(
                model="gpt-4",
                messages=[{"role": "user", "content": prompt}]
            )
            return response['choices'][0]['message']['content']
        return {
            "company_url": url,
            "error": f"Error fetching data: {str(e)}"
        }
    except Exception as e:
        return {
            "company_url": url,
            "error": f"Error fetching data: {str(e)}"
        }



def main():
    file_path = "sample.xlsx"
    # company_names = read_company_names(file_path)
    driver = setup_driver()
    scraped_data = []
    
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active
        company_names = []
        for row in sheet.iter_rows():
            if row[0].value == "Company Name":
                continue
            company_name = row[0].value
        # for company_name in company_names:
            print(f"Searching for: {company_name}")
            if company_name is None:
                continue
            company_data = {
                "name": company_name,
                "related_business": []
            }
            try:
                search_google_maps(driver, company_name)
                feed = driver.find_element(By.CSS_SELECTOR, "div[role='feed']")
                links = feed.find_elements(By.TAG_NAME, 'a')
                for link in links:
                    link_url = link.get_attribute('href')
                    if 'google.com/maps/place' not in link_url and 'skymem' not in link_url:
                        continue
                    # Open link in a new tab
                    original_window = driver.current_window_handle
                    driver.execute_script("window.open(arguments[0], '_blank');", link_url)
                    
                    # Wait for the new tab to load
                    WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))
                    time.sleep(3)
                    
                    # Switch to the new tab
                    for window_handle in driver.window_handles:
                        if window_handle != original_window:
                            driver.switch_to.window(window_handle)
                            break
                    
                    # Process data in the new tab
                    business_data = extract_related_business(driver)
                    business_data["google_map_link"] = link_url
                    company_data["related_business"].append(business_data)
                    driver.close()
                    driver.switch_to.window(original_window)
                    original_window = driver.current_window_handle
                    if business_data.get('url',None) is not None:
                        driver.execute_script("window.open(arguments[0], '_blank');", f'https://www.skymem.info/srch?q={business_data["url"]}')
                        WebDriverWait(driver, 10).until(EC.number_of_windows_to_be(2))
                        time.sleep(3)
                        for window_handle in driver.window_handles:
                            if window_handle != original_window:
                                driver.switch_to.window(window_handle)
                                break
                        try:
                            table = driver.find_element(By.TAG_NAME,"table")
                            emails = table.find_elements(By.TAG_NAME,"a")
                        except:
                            # driver.close()
                            emails = []
                            pass
                        emails_list = []
                        for email in emails :
                            emails_list.append(email.text)
                        # if len(emails_list) > 0:
                        driver.close()
                        business_data['emails'] = emails_list
                    driver.switch_to.window(original_window)
                    
            except :
                business_data = extract_related_business(driver)
                business_data["google_map_link"] = driver.current_url
                company_data["related_business"].append(business_data)

            scraped_data.append(company_data)
            save_to_json(scraped_data, "scraped_data.json")
            try:
                phones = [el.get('phone',None) for el in  company_data["related_business"] if el.get('phone',None) is not None]
                if len(phones) < 0:
                    phones = "غير موجود"
                else:
                    phones = "\n".join(phones)
            except:
                phones = "غير موجود"
            row[1].value = phones
            try:
                emails = {email for el in company_data["related_business"] for email in el.get('emails', [])}
                if len(emails) < 0:
                    emails = "غير موجود"
                else:
                    emails = "\n".join(emails)
            except:
                emails = "غير موجود"
            row[2].value = emails
            try:
                websites = [el.get('url',None) for el in  company_data["related_business"] if el.get('url',None) is not None]
                if len(websites) < 0:
                    websites = "غير موجود"
                else:
                    websites = "\n".join(websites)
            except:
                websites = "غير موجود"
            row[3].value = websites
            try:
                google_map_links = [el.get('google_map_link',None) for el in  company_data["related_business"]]
                if len(google_map_links) < 0:
                    google_map_links = "غير موجود"
                else:
                    google_map_links = "\n".join(google_map_links)
            except:
                google_map_links = "غير موجود"
            row[4].value = google_map_links
            # try:
            try:
                print("Processing social media accounts...")
                # Assuming `websites` is a string of newline-separated URLs
                websites = websites.split('\n')  # Replace `websites` with actual data
                social_media_results = collect_social_media_accounts("\n".join(websites))
                
                # Access dictionary keys directly
                for index in range(5, 11):
                    if index == 5:
                        facebooks = social_media_results.get("facebook", "")
                        facebooks = "\n".join(facebooks)
                        row[5].value = facebooks
                    elif index == 6:
                        linkedns = social_media_results.get("linkedin", "")
                        linkedns = "\n".join(linkedns)
                        row[6].value = linkedns
                    elif index == 7:
                        twitters = social_media_results.get("twitter", "")
                        twitters = "\n".join(twitters)
                        row[7].value = twitters
                    elif index == 8:
                        instagrams = social_media_results.get("instagram", "")
                        instagrams = "\n".join(instagrams)
                        row[8].value = instagrams
                    elif index == 9:
                        snapchats = social_media_results.get("snapchat", "")
                        snapchats = "\n".join(snapchats)
                        row[9].value = snapchats
                    else:
                        others = social_media_results.get("other", "")
                        others = "\n".join(others)
                        row[10].value = others
            except Exception as e:
                pass

            workbook.save(file_path)
            time.sleep(10)
            driver.get("https://www.google.com/maps")
    finally:
        driver.quit()

if __name__ == "__main__":
    main()
